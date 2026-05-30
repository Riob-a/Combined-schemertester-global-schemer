"""
generate_master_config.py
─────────────────────────
Generates master_config.xlsx from the existing per-conservancy Excel files.

Output: master_config.xlsx with two sheets:
  schemas           — one row per unique event_value, shared columns
  conservancy_mapping — one row per (event_value × conservancy),
                        with event_category_id and event_is_active

Run from the project root:
    python generate_master_config.py
    python generate_master_config.py --folder "excel files" --output master_config.xlsx
"""

import argparse
import glob
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# COLUMN DEFINITIONS
# ─────────────────────────────────────────────────────────────

# Columns that belong in the shared schemas sheet (one row per event_value).
# When the same event_value appears in multiple files and values conflict,
# the first conservancy in PRIORITY_ORDER wins.
SHARED_COLS = [
    "event_value",
    "schema_title",
    "api_version",
    "icon_id",
    "image_url",
    "n_fields",
    "n_enums",
    "event_name",
    "event_category_display",
]

# Columns that belong in the conservancy mapping sheet.
MAPPING_COLS = [
    "event_value",
    "conservancy",
    "event_category_id",
    "event_is_active",
]

# Tiebreaker priority when shared columns conflict.
# Earlier = higher priority.
PRIORITY_ORDER = ["borana", "mugie", "sosian", "suiyan", "catalyse", "test"]


# ─────────────────────────────────────────────────────────────
# LOAD ALL CONSERVANCY FILES
# ─────────────────────────────────────────────────────────────

def load_conservancy_files(folder: Path) -> dict[str, pd.DataFrame]:
    """Load all schemas-*.xlsx files from folder. Returns {conservancy: df}."""
    dfs = {}
    for f in sorted(glob.glob(str(folder / "schemas-*.xlsx"))):
        name = os.path.splitext(os.path.basename(f))[0].replace("schemas-", "")
        df = pd.read_excel(f, dtype={"event_value": str, "event_category_id": str})
        df["_conservancy"] = name
        dfs[name] = df
    if not dfs:
        raise FileNotFoundError(f"No schemas-*.xlsx files found in {folder}")
    return dfs


# ─────────────────────────────────────────────────────────────
# BUILD SCHEMAS SHEET
# ─────────────────────────────────────────────────────────────

def build_schemas_sheet(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    One row per unique event_value.
    When shared columns conflict, the conservancy earliest in PRIORITY_ORDER wins.
    Conservancies not in PRIORITY_ORDER are appended at the end.
    """
    priority = {name: i for i, name in enumerate(PRIORITY_ORDER)}
    ordered = sorted(dfs.keys(), key=lambda n: priority.get(n, 999))

    rows: dict[str, dict] = {}  # event_value → row dict

    for conservancy in ordered:
        df = dfs[conservancy]
        for _, row in df.iterrows():
            ev = str(row.get("event_value") or "").strip()
            if not ev or ev.lower() == "nan":
                continue

            if ev not in rows:
                # First time seeing this event_value — take all shared cols
                entry = {}
                for col in SHARED_COLS:
                    val = row.get(col)
                    entry[col] = val if pd.notna(val) else None
                rows[ev] = entry
            else:
                # Already seen — fill in any blanks from this conservancy
                for col in SHARED_COLS:
                    if col == "event_value":
                        continue
                    if rows[ev].get(col) is None:
                        val = row.get(col)
                        if pd.notna(val):
                            rows[ev][col] = val

    df_out = pd.DataFrame(list(rows.values()), columns=SHARED_COLS)
    return df_out.sort_values("event_value").reset_index(drop=True)


# ─────────────────────────────────────────────────────────────
# BUILD CONSERVANCY MAPPING SHEET
# ─────────────────────────────────────────────────────────────

def build_mapping_sheet(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    One row per (event_value × conservancy).
    Tracks event_category_id and event_is_active per conservancy.
    """
    rows = []
    for conservancy, df in dfs.items():
        for _, row in df.iterrows():
            ev = str(row.get("event_value") or "").strip()
            if not ev or ev.lower() == "nan":
                continue
            cat_id = row.get("event_category_id")
            is_active = row.get("event_is_active")
            rows.append({
                "event_value":      ev,
                "conservancy":      conservancy,
                "event_category_id": str(cat_id).strip() if pd.notna(cat_id) else None,
                "event_is_active":  is_active if pd.notna(is_active) else None,
            })

    df_out = pd.DataFrame(rows, columns=MAPPING_COLS)
    return df_out.sort_values(["event_value", "conservancy"]).reset_index(drop=True)


# ─────────────────────────────────────────────────────────────
# FORMATTING
# ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")  # dark blue
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=10)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header formatting and auto-width columns."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    # Auto-width: max of header length and longest value, capped at 60
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        values = df.iloc[:, col_idx - 1].astype(str).tolist() + [col_name]
        width  = min(max(len(str(v)) for v in values) + 2, 60)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────
# WRITE OUTPUT
# ─────────────────────────────────────────────────────────────

def write_master_config(
    df_schemas: pd.DataFrame,
    df_mapping: pd.DataFrame,
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_schemas.to_excel(writer, sheet_name="schemas", index=False)
        df_mapping.to_excel(writer, sheet_name="conservancy_mapping", index=False)

    # Apply formatting via openpyxl after pandas writes the data
    wb = load_workbook(output_path)
    _format_sheet(wb["schemas"],             df_schemas)
    _format_sheet(wb["conservancy_mapping"], df_mapping)
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────
# CONFLICT REPORT
# ─────────────────────────────────────────────────────────────

def print_conflict_report(dfs: dict[str, pd.DataFrame]) -> None:
    """
    Report event_values where shared columns differ across conservancies.
    Informational only — the priority order resolves all conflicts automatically.
    """
    combined = pd.concat(dfs.values(), ignore_index=True)
    counts   = combined.groupby("event_value").size()
    multi    = counts[counts > 1].index

    check_cols = ["schema_title", "icon_id", "event_name", "event_category_display"]
    conflicts  = []

    for ev in multi:
        rows = combined[combined["event_value"] == ev]
        for col in check_cols:
            if col not in rows.columns:
                continue
            unique_vals = rows[col].dropna().unique()
            if len(unique_vals) > 1:
                conservancies = rows.dropna(subset=[col])["_conservancy"].tolist()
                conflicts.append({
                    "event_value":   ev,
                    "column":        col,
                    "values":        list(unique_vals),
                    "conservancies": conservancies,
                })

    if not conflicts:
        print("  No conflicts in shared columns.")
        return

    print(f"  {len(conflicts)} conflict(s) resolved by priority order ({', '.join(PRIORITY_ORDER)}):")
    for c in conflicts:
        print(f"    {c['event_value']} | {c['column']}")
        for val, con in zip(c["values"], c["conservancies"]):
            print(f"      {con}: {val!r}")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate master_config.xlsx")
    parser.add_argument(
        "--folder", type=str, default="excel files",
        help="Folder containing schemas-*.xlsx files (default: 'excel files')",
    )
    parser.add_argument(
        "--output", type=str, default="master_config.xlsx",
        help="Output file path (default: master_config.xlsx)",
    )
    args = parser.parse_args()

    folder      = Path(args.folder)
    output_path = Path(args.output)

    print(f"Loading conservancy files from: {folder}")
    dfs = load_conservancy_files(folder)
    print(f"  Loaded {len(dfs)} conservancy files: {', '.join(sorted(dfs.keys()))}")

    print("\nConflict report (shared columns):")
    print_conflict_report(dfs)

    print("\nBuilding sheets...")
    df_schemas = build_schemas_sheet(dfs)
    df_mapping = build_mapping_sheet(dfs)

    print(f"  schemas sheet:             {len(df_schemas)} rows ({df_schemas['event_value'].nunique()} unique event_values)")
    print(f"  conservancy_mapping sheet: {len(df_mapping)} rows")

    print(f"\nWriting {output_path} ...")
    write_master_config(df_schemas, df_mapping, output_path)
    print(f"  Done.")


if __name__ == "__main__":
    main()
